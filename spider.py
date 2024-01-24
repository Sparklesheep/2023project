# -*- coding = utf-8 -*-
"""
@File: spider.py
@Author : Tianci
@Time : 2024/1/23 14:15
@Software : PyCharm
"""

import requests
from requests.auth import HTTPBasicAuth
from datetime import datetime
import time
import openpyxl
from login import info


def get_data(username, repo_name, path, max_pages, next_page_url=None):
    base_url = f'https://api.github.com/repos/{username}/{repo_name}/commits'

    username = info.username
    password = info.password

    headers = {
        'User-Agent': username
    }

    if next_page_url is None:
        response = requests.get(base_url, headers=headers, auth=HTTPBasicAuth(username, password))
    else:
        response = requests.get(next_page_url, headers=headers, auth=HTTPBasicAuth(username, password))

    datalist = []
    counter = 0

    if response.status_code == 200:
        commits = response.json()
        for commit in commits:

            counter += 1
            data = []
            commit_sha = commit['sha']
            commit_message = commit['commit']['message']
            commit_author = commit['commit']['author']['name']
            commit_username = commit['author']['login'] if commit['author'] is not None and 'login' in commit['author'] else 'Unknown'
            commit_date_time = commit['commit']['author']['date']
            commit_comment = commit['commit']['comment_count']

            print('counter = ' + str(counter))
            print(f"SHA: {commit_sha}")
            print(f"Author: {commit_author}")
            print(f"Username: {commit_username}")
            print(f"Date: {commit_date_time}")
            print(f"Message: {commit_message}")
            print(f"comment_count: {commit_comment}")
            print("\n" + "=" * 50 + "\n")

            date_time_obj = datetime.strptime(commit_date_time, "%Y-%m-%dT%H:%M:%SZ")

            commit_date = date_time_obj.date()
            commit_time = date_time_obj.time()

            data.append(commit_sha)
            data.append(commit_username)
            data.append(commit_author)
            data.append(commit_date)
            data.append(commit_time)
            data.append(commit_message)
            data.append(commit_comment)
            datalist.append(data)

        save_data(datalist, path)

        time.sleep(2)
        max_pages = max_pages - 1
        if 'Link' in response.headers:
            links = response.headers['Link'].split(', ')
            for link in links:
                if 'rel="next"' in link and max_pages > 0:
                    next_page_url = link.split(';')[0][1:-1]
                    print(next_page_url)
                    print(max_pages)
                    return get_data(username, repo_name, path, max_pages, next_page_url)
        print(datalist)

    else:
        print(f"Error: {response.status_code}")
        print(f"Error Message: {response.text}")

    return datalist


def init_data(path):
    workbook = openpyxl.load_workbook(path)
    workbook.remove(workbook['Sheet1'])
    workbook.create_sheet('Sheet1')
    sheet = workbook.active
    sheet.append(['sha', 'username', 'author', 'date', 'time', 'message', 'comment'])
    workbook.save(path)


def save_data(datalist, path):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook.active
    counter = 0
    for row in datalist:
        counter += 1
        sheet.append(row)
    workbook.save(path)


if __name__ == "__main__":
    excel_path = 'commits.xlsx'
    init_data(excel_path)
    data = get_data('sveltejs', 'svelte', excel_path, max_pages=40)

